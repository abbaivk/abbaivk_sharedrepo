from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


COLORS = {
    "header": "1F4E78",
    "h2": "245B89",
    "top": "D9EAF7",
    "current": "EAF5F1",
    "previous": "FFF4D8",
    "date": "EEF2F6",
    "white": "FFFFFF",
    "alternate": "F7F9FB",
    "text": "1F2933",
    "muted": "52616F",
    "green_fill": "E8F5E9",
    "green_text": "1B6B35",
    "amber_fill": "FFF4D8",
    "amber_text": "8A5A00",
    "red_fill": "FDECEC",
    "red_text": "A12622",
    "blue_fill": "E6F0FF",
    "blue_text": "1F4E78",
    "gray_fill": "EEF2F6",
    "gray_text": "52616F",
}


def fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def cell_text(cell) -> str:
    return "" if cell.value is None else str(cell.value).strip()


def row_values(ws, row: int) -> list[str]:
    return [cell_text(ws.cell(row=row, column=col)) for col in range(1, ws.max_column + 1)]


def is_blank_row(values: list[str]) -> bool:
    return not any(values)


def status_style(status: str) -> tuple[str, str]:
    value = status.lower()
    if "off track" in value or "delayed" in value:
        return COLORS["red_fill"], COLORS["red_text"]
    if "risk" in value:
        return COLORS["amber_fill"], COLORS["amber_text"]
    if "complete" in value or "done" in value:
        return COLORS["blue_fill"], COLORS["blue_text"]
    if "track" in value:
        return COLORS["green_fill"], COLORS["green_text"]
    return COLORS["gray_fill"], COLORS["gray_text"]


def header_map(values: list[str]) -> dict[str, int]:
    return {value: index + 1 for index, value in enumerate(values) if value}


def first_present(headers: dict[str, int], names: list[str]) -> int | None:
    for name in names:
        if name in headers:
            return headers[name]
    return None


def looks_like_header(values: list[str]) -> bool:
    labels = {
        "H2 Order",
        "H2 Value",
        "Top-Level Method",
        "Sub-Level Method",
        "Sub-Method",
        "Current Status",
        "Method Level",
        "Method",
        "Status",
    }
    return len(labels.intersection(values)) >= 2


def style_row(ws, row: int, last_col: int, fill_color: str, font_color: str, bold: bool = False, size: int = 12):
    for col in range(1, last_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill(fill_color)
        cell.font = Font(color=font_color, bold=bold, size=size)
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def style_header(ws, row: int, last_col: int):
    style_row(ws, row, last_col, COLORS["header"], "FFFFFF", bold=True, size=13)
    for col in range(1, last_col + 1):
        ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_status_cell(cell):
    fill_color, text_color = status_style(cell_text(cell))
    cell.fill = fill(fill_color)
    cell.font = Font(color=text_color, bold=True, size=12)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_sheet(ws):
    last_col = max(ws.max_column, 1)
    header_rows: list[tuple[int, dict[str, int]]] = []

    for row in range(1, ws.max_row + 1):
        values = row_values(ws, row)
        if looks_like_header(values):
            style_header(ws, row, last_col)
            header_rows.append((row, header_map(values)))
        elif row == 1 and any(values):
            ws.cell(row=1, column=1).font = Font(color=COLORS["blue_text"], bold=True, size=14)
            ws.cell(row=1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        elif values[0] in {"Report Summary", "H2 Value Summary"}:
            style_row(ws, row, last_col, COLORS["header"], "FFFFFF", bold=True, size=13)

    if not header_rows:
        return

    for header_row, headers in header_rows:
        current_start = first_present(headers, ["Current Status", "Status"])
        current_end = first_present(headers, ["Current Commentary", "Commentary"])
        previous_start = first_present(headers, ["Previous Status"])
        previous_end = first_present(headers, ["Previous Commentary"])
        status_cols = [
            first_present(headers, ["Current Status", "Status"]),
            first_present(headers, ["Previous Status"]),
        ]
        date_cols = [
            first_present(headers, ["Status Update Date"]),
            first_present(headers, ["Update Date"]),
            first_present(headers, ["Comment Date"]),
        ]
        method_cols = [
            first_present(headers, ["Top-Level Method", "Method"]),
            first_present(headers, ["Sub-Level Method", "Sub-Method"]),
        ]
        h2_order_col = first_present(headers, ["H2 Order", "Order"])
        h2_value_col = first_present(headers, ["H2 Value"])

        detail_index = 0
        for row in range(header_row + 1, ws.max_row + 1):
            values = row_values(ws, row)
            if is_blank_row(values):
                continue
            if looks_like_header(values):
                break

            non_empty = [value for value in values if value]
            if (
                h2_order_col
                and h2_value_col
                and cell_text(ws.cell(row=row, column=h2_order_col))
                and cell_text(ws.cell(row=row, column=h2_value_col))
                and len(non_empty) <= 2
            ):
                style_row(ws, row, last_col, COLORS["h2"], "FFFFFF", bold=True, size=13)
                continue

            if len(non_empty) == 1 and any(value.startswith("Top-Level Method:") for value in values):
                style_row(ws, row, last_col, COLORS["top"], "173A56", bold=True)
                continue

            base_fill = COLORS["white"] if detail_index % 2 == 0 else COLORS["alternate"]
            style_row(ws, row, last_col, base_fill, COLORS["text"])
            detail_index += 1

            if current_start and current_end and current_end >= current_start:
                for col in range(current_start, current_end + 1):
                    ws.cell(row=row, column=col).fill = fill(COLORS["current"])

            if previous_start and previous_end and previous_end >= previous_start:
                for col in range(previous_start, previous_end + 1):
                    ws.cell(row=row, column=col).fill = fill(COLORS["previous"])

            for col in [col for col in date_cols if col]:
                cell = ws.cell(row=row, column=col)
                cell.fill = fill(COLORS["date"])
                cell.font = Font(color=COLORS["muted"], size=12)

            for col in [col for col in status_cols if col]:
                style_status_cell(ws.cell(row=row, column=col))

            for col in [col for col in method_cols if col]:
                cell = ws.cell(row=row, column=col)
                cell.font = Font(color=COLORS["text"], bold=True, size=12)

    ws.freeze_panes = "A2"
    for col in range(1, last_col + 1):
        letter = get_column_letter(col)
        if ws.column_dimensions[letter].width is None:
            ws.column_dimensions[letter].width = 18


def style_workbook(path: Path):
    wb = load_workbook(path)
    for ws in wb.worksheets:
        style_sheet(ws)
    wb.save(path)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        raise SystemExit("Pass one or more .xlsx files")
    for arg in sys.argv[1:]:
        style_workbook(Path(arg))
    print("styled", len(sys.argv) - 1, "workbook(s)")
